import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
# for printing sheet names
# print(wb.sheetnames)

# selecting a sheet
# sheet = wb['Sheet1']
# getting the active sheet
sheet = wb.active

# creating and removing worksheets with index
# wb.create_sheet('Sheet2', 0)
# wb.remove(wb["Sheet2"])
print(wb.sheetnames)

# Accessing an individual cell or a range of cells
# passing coordinate of a cell
# cell = sheet['A1'] 
# changing the value of a cell
# cell.value = "id"
# print(cell.value)
# printing row, column and coordinate info
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# accessing cell with sheet.cell method with keyword args
# cell = sheet.cell(row=1, column=1)
# print(cell.value)

# iterating over all rows and columns
# for row in range(1, sheet.max_row +1):
#   for column in range(1, sheet.max_column +1):
#     cell = sheet.cell(row, column)
#     print(cell.value)

# accessing a range of cells with sheet[] 
# accessing all cells of a column
# columnA = sheet['a']
# for cell in columnA:
#   print(cell.value)

# accessing all cells of a row
# rowA = sheet['1']
# for cell in rowA:
#   print(cell.value)


# accessing a specific range of cells
# accessing all cells from column A to column C
# cellsByRow = sheet['A:C']
# print(cellsByRow)

# accessing all cells from row a to row 4
# cellsByRow = sheet['1:4'] # or sheet[1:4] # without quotes
# print(cellsByRow)

# accessing all cells from one cordinate to another cordinate
# cellsByRow = sheet['a1:c3']
# print(cellsByRow)


# Some useful method of sheet object
# Adding a row with append method
# sheet.append([1,2,3])

# Inserting empty rows at the given index
# sheet.insert_rows(idx=1, amount=1)

# moving a range of cells
# sheet.move_range(cell_range="A2:C5", rows=-1, cols=0)


wb.save('transactions.xlsx')